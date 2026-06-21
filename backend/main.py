import csv
import base64
import io
import re
import secrets
import zipfile
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse, Response
from pydantic import BaseModel
import cloudinary
import cloudinary.uploader

from . import config, storage, renderer, senders, jobs
from .senders.base import Recipient

app = FastAPI(title="전인교육학회 뉴스레터 API")
CONSENT_VERSION = "newsletter-consent-v1-2026-05-28"

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

PUBLIC_PATHS = {"/api/subscribe", "/api/unsubscribe", "/health"}


def _is_public_path(path: str) -> bool:
    normalized = path.rstrip("/") or "/"
    return normalized in PUBLIC_PATHS


def _authorized(auth_header: str | None) -> bool:
    if not config.ADMIN_PASSWORD:
        return True
    if not auth_header or not auth_header.startswith("Basic "):
        return False
    try:
        raw = base64.b64decode(auth_header.removeprefix("Basic ").strip()).decode("utf-8")
        username, password = raw.split(":", 1)
    except Exception:
        return False
    return (
        secrets.compare_digest(username.encode("utf-8"), config.ADMIN_USERNAME.encode("utf-8"))
        and secrets.compare_digest(password.encode("utf-8"), config.ADMIN_PASSWORD.encode("utf-8"))
    )


@app.middleware("http")
async def require_admin_auth(request: Request, call_next):
    if config.ADMIN_PASSWORD and not _is_public_path(request.url.path):
        if not _authorized(request.headers.get("authorization")):
            return PlainTextResponse(
                "Authentication required",
                status_code=401,
                headers={"WWW-Authenticate": 'Basic realm="Newsletter Admin"'},
            )
    return await call_next(request)

if config.CLOUDINARY_CLOUD_NAME:
    cloudinary.config(
        cloud_name=config.CLOUDINARY_CLOUD_NAME,
        api_key=config.CLOUDINARY_API_KEY,
        api_secret=config.CLOUDINARY_API_SECRET,
        secure=True,
    )


class RenderRequest(BaseModel):
    template: str = "classic"
    data: dict


class DraftIn(BaseModel):
    id: str | None = None
    name: str
    template: str
    subject: str
    data: dict


class SendRequest(BaseModel):
    sender: str
    template: str
    subject: str
    from_addr: str
    from_name: str = "전인교육학회"
    recipients: list[dict]  # [{email, name?}]
    data: dict


@app.get("/api/templates")
def get_templates():
    return [{"key": k, **v} for k, v in renderer.AVAILABLE_TEMPLATES.items()]


@app.get("/api/senders")
def get_senders():
    return senders.get_available()


@app.post("/api/render", response_class=HTMLResponse)
def post_render(req: RenderRequest):
    try:
        return renderer.render(req.template, req.data)
    except Exception as e:
        raise HTTPException(400, str(e))


@app.post("/api/export")
def export_html(req: RenderRequest):
    html = renderer.render(req.template, req.data)
    return Response(
        content=html,
        media_type="text/html",
        headers={"Content-Disposition": 'attachment; filename="newsletter.html"'},
    )


@app.get("/api/drafts")
def get_drafts():
    return storage.list_drafts()


@app.get("/api/drafts/{draft_id}")
def get_draft(draft_id: str):
    d = storage.get_draft(draft_id)
    if not d:
        raise HTTPException(404, "Not found")
    return d


@app.post("/api/drafts")
def post_draft(d: DraftIn):
    return storage.save_draft(d.model_dump())


@app.delete("/api/drafts/{draft_id}")
def delete_draft(draft_id: str):
    return {"ok": storage.delete_draft(draft_id)}


class RenameIn(BaseModel):
    name: str


@app.patch("/api/drafts/{draft_id}")
def rename_draft(draft_id: str, body: RenameIn):
    d = storage.rename_draft(draft_id, body.name)
    if not d:
        raise HTTPException(404, "Not found")
    return d


@app.get("/api/default-draft")
def get_default_draft():
    return {"id": storage.get_default_draft_id()}


class DefaultDraftIn(BaseModel):
    id: str | None = None


@app.put("/api/default-draft")
def put_default_draft(body: DefaultDraftIn):
    ok = storage.set_default_draft_id(body.id)
    if not ok:
        raise HTTPException(404, "Draft not found")
    return {"ok": True, "id": body.id}


@app.post("/api/upload-image")
async def upload_image(file: UploadFile = File(...)):
    if not config.CLOUDINARY_CLOUD_NAME:
        raise HTTPException(500, "Cloudinary 환경변수가 설정되지 않았습니다 (.env 확인)")
    content = await file.read()
    res = cloudinary.uploader.upload(content, folder="newsletter")
    return {"url": res["secure_url"], "public_id": res["public_id"]}


EMAIL_RE = re.compile(r"^[^@\s<>;,]+@[^@\s<>;,]+\.[^@\s<>;,]+$")
FILTER_INPUT_EXTENSIONS = (".csv", ".txt", ".xlsx", ".xlsm")
FILTER_ARCHIVE_EXTENSIONS = (".zip",)
FILTER_RESPONSE_RECIPIENT_LIMIT = 50000
EMAIL_HEADER_KEYS = {
    "email",
    "emailaddress",
    "mail",
    "메일",
    "메일주소",
    "이메일",
    "이메일주소",
    "전자메일",
    "전자우편",
    "수신이메일",
    "수신자이메일",
}
NAME_HEADER_KEYS = {"name", "fullname", "recipient", "recipientname", "이름", "성명", "성함", "수신자", "받는사람"}


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value) -> str:
    return re.sub(r"[\s_\-().]+", "", _clean_cell(value).lower())


def _normalize_email(value) -> str:
    return _clean_cell(value).lower()


def _is_valid_email(value) -> bool:
    return bool(EMAIL_RE.match(_normalize_email(value)))


def _unique_columns(raw_header: list[str], width: int) -> list[str]:
    columns: list[str] = []
    seen: dict[str, int] = {}
    for i in range(width):
        base = _clean_cell(raw_header[i]) if i < len(raw_header) else ""
        if not base:
            base = f"column_{i + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        columns.append(base if count == 0 else f"{base}_{count + 1}")
    return columns


def _decode_table_text(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            return content.decode("cp949")
        except UnicodeDecodeError:
            return content.decode("utf-8", errors="replace")


def _table_rows_from_upload(filename: str, content: bytes) -> list[list[str]]:
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows = [[_clean_cell(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        except Exception as e:
            raise HTTPException(400, f"Excel 파싱 실패: {e}")
    else:
        raw = _decode_table_text(content)
        try:
            dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        rows = [[_clean_cell(cell) for cell in row] for row in csv.reader(io.StringIO(raw), dialect)]
    return [row for row in rows if any(cell for cell in row)]


def _zip_supported_members(content: bytes) -> list[tuple[str, bytes]]:
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile:
        raise HTTPException(400, "ZIP 파일을 읽을 수 없습니다.")

    files: list[tuple[str, bytes]] = []
    for info in archive.infolist():
        member_name = info.filename
        base_name = Path(member_name).name
        if info.is_dir() or not base_name or base_name.startswith(".") or member_name.startswith("__MACOSX/"):
            continue
        if not member_name.lower().endswith(FILTER_INPUT_EXTENSIONS):
            continue
        files.append((member_name, archive.read(info)))
    return files


def _header_email_index(row: list[str]) -> int | None:
    for i, value in enumerate(row):
        if _normalize_header(value) in EMAIL_HEADER_KEYS:
            return i
    return None


def _header_name_index(columns: list[str]) -> int | None:
    for i, value in enumerate(columns):
        if _normalize_header(value) in NAME_HEADER_KEYS:
            return i
    return None


def _best_email_column(rows: list[list[str]], width: int) -> int | None:
    best_i = None
    best_count = 0
    for i in range(width):
        count = sum(1 for row in rows if i < len(row) and _is_valid_email(row[i]))
        if count > best_count:
            best_i = i
            best_count = count
    return best_i if best_count else None


def _tabular_records(rows: list[list[str]]) -> tuple[list[str], list[dict], int | None, int | None]:
    if not rows:
        return [], [], None, None
    width = max(len(row) for row in rows)
    first = rows[0]
    header_email_i = _header_email_index(first)

    if header_email_i is not None:
        columns = _unique_columns(first, width)
        data_rows = rows[1:]
        email_i = header_email_i
    else:
        first_has_email = any(_is_valid_email(cell) for cell in first)
        email_i_after_header = _best_email_column(rows[1:], width) if len(rows) > 1 else None
        if not first_has_email and email_i_after_header is not None:
            columns = _unique_columns(first, width)
            data_rows = rows[1:]
            email_i = email_i_after_header
        else:
            columns = [f"column_{i + 1}" for i in range(width)]
            data_rows = rows
            email_i = _best_email_column(rows, width)

    name_i = _header_name_index(columns)
    records = []
    for row in data_rows:
        values = {columns[i]: (_clean_cell(row[i]) if i < len(row) else "") for i in range(width)}
        email_raw = _clean_cell(row[email_i]) if email_i is not None and email_i < len(row) else ""
        name = _clean_cell(row[name_i]) if name_i is not None and name_i < len(row) else None
        records.append({"values": values, "email_raw": email_raw, "email": _normalize_email(email_raw), "name": name or None})
    return columns, records, email_i, name_i


def _filter_recipient_bytes(filename: str, content: bytes, unsubscribed: set[str]) -> dict:
    rows = _table_rows_from_upload(filename.lower(), content)
    columns, records, email_i, _ = _tabular_records(rows)
    seen: set[str] = set()
    kept = []
    removed_unsubscribed = []
    removed_duplicate = []
    removed_invalid = []

    for record in records:
        email = record["email"]
        public_record = {"values": record["values"], "email": email, "name": record["name"]}
        if email_i is None or not _is_valid_email(record["email_raw"]):
            removed_invalid.append({**public_record, "reason": "invalid_email"})
            continue
        if email in seen:
            removed_duplicate.append({**public_record, "reason": "duplicate"})
            continue
        seen.add(email)
        if email in unsubscribed:
            removed_unsubscribed.append({**public_record, "reason": "unsubscribed"})
            continue
        kept.append(public_record)

    return {
        "filename": filename,
        "columns": columns,
        "kept": kept,
        "removed_unsubscribed": removed_unsubscribed,
        "removed_duplicate": removed_duplicate,
        "removed_invalid": removed_invalid,
        "counts": {
            "total": len(records),
            "kept": len(kept),
            "unsubscribed": len(removed_unsubscribed),
            "duplicate": len(removed_duplicate),
            "invalid": len(removed_invalid),
        },
    }


def _load_unsubscribed_or_error() -> set[str]:
    try:
        return storage.load_unsubscribed()
    except Exception as exc:
        raise HTTPException(
            503,
            f"수신거부 명단을 불러오지 못했습니다. Supabase 환경변수와 서비스 키를 확인해 주세요. ({exc})",
        )


def _aggregate_filter_results(files: list[dict]) -> dict:
    counts = {"total": 0, "kept": 0, "unsubscribed": 0, "duplicate": 0, "invalid": 0}
    for result in files:
        for key in counts:
            counts[key] += result.get("counts", {}).get(key, 0)
    return {
        "counts": counts,
        "kept": [row for result in files for row in result.get("kept", [])],
        "removed_unsubscribed": [row for result in files for row in result.get("removed_unsubscribed", [])],
        "removed_duplicate": [row for result in files for row in result.get("removed_duplicate", [])],
        "removed_invalid": [row for result in files for row in result.get("removed_invalid", [])],
    }


def _filter_upload(filename: str, content: bytes) -> dict:
    lower = filename.lower()
    supported = (*FILTER_INPUT_EXTENSIONS, *FILTER_ARCHIVE_EXTENSIONS)
    if not lower.endswith(supported):
        raise HTTPException(400, "지원 포맷: CSV, TXT, Excel(.xlsx/.xlsm), ZIP")

    unsubscribed = _load_unsubscribed_or_error()
    if lower.endswith(FILTER_ARCHIVE_EXTENSIONS):
        members = _zip_supported_members(content)
        if not members:
            raise HTTPException(400, "ZIP 내부에 지원되는 CSV, TXT, Excel(.xlsx/.xlsm) 파일이 없습니다.")
        files = []
        errors = []
        for name, member_content in members:
            try:
                files.append(_filter_recipient_bytes(name, member_content, unsubscribed))
            except HTTPException as exc:
                errors.append({"filename": name, "detail": exc.detail})
            except Exception as exc:
                errors.append({"filename": name, "detail": str(exc)})
        if not files:
            first_error = errors[0]["detail"] if errors else "처리 가능한 파일이 없습니다."
            raise HTTPException(400, f"ZIP 내부 파일 처리 실패: {first_error}")
        aggregate = _aggregate_filter_results(files)
        return {
            "filename": filename,
            "is_zip": True,
            "file_count": len(files),
            "error_count": len(errors),
            "errors": errors,
            "files": files,
            "columns": files[0]["columns"] if files else [],
            **aggregate,
        }

    result = _filter_recipient_bytes(filename, content, unsubscribed)
    return {
        "filename": filename,
        "is_zip": False,
        "file_count": 1,
        "error_count": 0,
        "errors": [],
        "files": [result],
        **result,
    }


def _compact_filter_result(result: dict) -> dict:
    kept = result.get("kept", [])
    kept_for_response = kept[:FILTER_RESPONSE_RECIPIENT_LIMIT]
    kept_truncated = len(kept) > len(kept_for_response)
    return {
        "filename": result.get("filename"),
        "is_zip": result.get("is_zip", False),
        "file_count": result.get("file_count", 1),
        "error_count": result.get("error_count", 0),
        "errors": result.get("errors", []),
        "columns": result.get("columns", []),
        "counts": result.get("counts", {}),
        "kept": [{"email": row.get("email"), "name": row.get("name")} for row in kept_for_response],
        "kept_truncated": kept_truncated,
        "recipient_limit": FILTER_RESPONSE_RECIPIENT_LIMIT,
        "can_use_as_recipients": not kept_truncated,
        "files": [
            {
                "filename": file_result.get("filename"),
                "columns": file_result.get("columns", []),
                "counts": file_result.get("counts", {}),
            }
            for file_result in result.get("files", [])
        ],
    }


def _csv_bytes_for_filter_result(result: dict, kind: str) -> bytes:
    columns = result.get("columns", [])
    if kind == "kept":
        header = columns
        records = result.get("kept", [])
        rows = [[record.get("values", {}).get(column, "") for column in columns] for record in records]
    else:
        header = [*columns, "제외 사유", "판별 이메일"]
        reason_groups = [
            ("수신거부", result.get("removed_unsubscribed", [])),
            ("중복", result.get("removed_duplicate", [])),
            ("이메일 오류", result.get("removed_invalid", [])),
        ]
        rows = []
        for label, records in reason_groups:
            for record in records:
                rows.append([
                    *[record.get("values", {}).get(column, "") for column in columns],
                    label,
                    record.get("email", ""),
                ])

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _safe_csv_name(filename: str, suffix: str) -> str:
    stem = Path(filename).stem or "recipients"
    stem = re.sub(r"[^\w가-힣.-]+", "_", stem).strip("._") or "recipients"
    return f"{stem}_{suffix}.csv"


def _unique_zip_csv_name(filename: str, suffix: str, used_names: set[str]) -> str:
    candidate = _safe_csv_name(filename, suffix)
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    stem = Path(candidate).stem
    ext = Path(candidate).suffix
    index = 2
    while True:
        indexed = f"{stem}_{index}{ext}"
        if indexed not in used_names:
            used_names.add(indexed)
            return indexed
        index += 1


@app.post("/api/filter-recipients")
async def filter_recipients(file: UploadFile = File(...)):
    return _compact_filter_result(_filter_upload(file.filename or "recipients", await file.read()))


@app.post("/api/filter-recipients/export")
async def export_filtered_recipients(file: UploadFile = File(...), kind: str = Form("kept")):
    if kind not in {"kept", "excluded"}:
        raise HTTPException(400, "kind는 kept 또는 excluded만 지원합니다.")

    result = _filter_upload(file.filename or "recipients", await file.read())
    suffix = "filtered" if kind == "kept" else "excluded"
    if result["is_zip"]:
        buffer = io.BytesIO()
        used_names: set[str] = set()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file_result in result["files"]:
                archive.writestr(_unique_zip_csv_name(file_result["filename"], suffix, used_names), _csv_bytes_for_filter_result(file_result, kind))
            if result.get("errors"):
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(["파일명", "오류"])
                for error in result["errors"]:
                    writer.writerow([error.get("filename", ""), error.get("detail", "")])
                archive.writestr("_processing_errors.csv", ("\ufeff" + output.getvalue()).encode("utf-8"))
        return Response(
            content=buffer.getvalue(),
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="unsubscribe_{suffix}_files.zip"'},
        )

    return Response(
        content=_csv_bytes_for_filter_result(result["files"][0], kind),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="unsubscribe_{suffix}_recipients.csv"'},
    )


@app.post("/api/parse-recipients")
async def parse_recipients(file: UploadFile = File(...)):
    """CSV/XLSX 파싱: 'email' 단일 컬럼 또는 'email,name' 모두 지원."""
    filename = (file.filename or "").lower()
    content = await file.read()

    # Excel (.xlsx) 처리
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"recipients": [], "count": 0}
            # 첫 행을 헤더로 시도
            header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
            out = []
            if "email" in header:
                email_i = header.index("email")
                name_i = header.index("name") if "name" in header else None
                for row in rows[1:]:
                    if not row or email_i >= len(row):
                        continue
                    email = (str(row[email_i]).strip() if row[email_i] else "")
                    if not email or "@" not in email:
                        continue
                    name = None
                    if name_i is not None and name_i < len(row) and row[name_i]:
                        name = str(row[name_i]).strip()
                    out.append({"email": email, "name": name})
            else:
                # 헤더 없이 첫 컬럼을 이메일로 취급
                for row in rows:
                    if not row:
                        continue
                    cell = row[0]
                    if cell and "@" in str(cell):
                        out.append({"email": str(cell).strip()})
            return {"recipients": out, "count": len(out)}
        except Exception as e:
            raise HTTPException(400, f"Excel 파싱 실패: {e}")

    # CSV 처리
    try:
        raw = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            raw = content.decode("cp949")  # 한국어 엑셀에서 자주 보임
        except UnicodeDecodeError:
            raw = content.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(raw))
    fields = [f.strip().lower() for f in (reader.fieldnames or [])]
    if "email" not in fields:
        out = []
        for line in raw.splitlines():
            line = line.strip()
            if line and "@" in line:
                out.append({"email": line})
        return {"recipients": out, "count": len(out)}
    out = []
    for row in reader:
        norm = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
        if norm.get("email"):
            out.append({"email": norm["email"], "name": norm.get("name") or None})
    return {"recipients": out, "count": len(out)}


def _filter_send_recipients(recipients: list[dict]) -> tuple[list[Recipient], int]:
    unsub = storage.load_unsubscribed()
    seen: set[str] = set()
    out: list[Recipient] = []
    skipped = 0
    for raw in recipients:
        email = (raw.get("email") or "").strip().lower()
        if not email or "@" not in email:
            continue
        if email in seen:
            continue
        seen.add(email)
        if email in unsub:
            skipped += 1
            continue
        out.append(Recipient(email=email, name=raw.get("name")))
    return out, skipped


def _split_batch_result(batch: list[dict], errors: list[str], failed_count: int) -> tuple[list[str], list[dict]]:
    by_email = {row["email"].strip().lower(): row for row in batch}
    failed_by_id: dict[str, dict] = {}
    for err in errors:
        prefix = err.split(":", 1)[0].strip().lower() if ":" in err else ""
        row = by_email.get(prefix)
        if row:
            failed_by_id[row["id"]] = {**row, "error": err}

    if failed_count and not failed_by_id:
        message = errors[0] if errors else "발송 실패"
        for row in batch[:failed_count]:
            failed_by_id[row["id"]] = {**row, "error": message}
    elif len(failed_by_id) < failed_count:
        message = errors[0] if errors else "발송 실패"
        remaining = [row for row in batch if row["id"] not in failed_by_id]
        for row in remaining[: failed_count - len(failed_by_id)]:
            failed_by_id[row["id"]] = {**row, "error": message}

    failed_ids = set(failed_by_id)
    sent_ids = [row["id"] for row in batch if row["id"] not in failed_ids]
    return sent_ids, list(failed_by_id.values())


@app.post("/api/send")
def post_send(req: SendRequest):
    renderer.render(req.template, req.data)
    rcpts, skipped = _filter_send_recipients(req.recipients)

    if not rcpts:
        return {"job_id": None, "sent": 0, "failed": 0, "skipped": skipped, "message": "발송할 수신자가 없습니다 (모두 수신거부됨)"}

    job = jobs.create_job(req.model_dump(), rcpts, skipped=skipped, batch_size=config.SEND_BATCH_SIZE)
    return {"job_id": job["id"], "total": job["total"], "skipped": job["skipped"]}


@app.post("/api/send/{job_id}/process")
def process_send_job(job_id: str):
    job = jobs.get_job(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    if job["status"] in {"done", "failed"}:
        return job

    batch = jobs.get_pending_recipients(job_id, config.SEND_BATCH_SIZE)
    if not batch:
        done = jobs.finish_if_complete(job_id)
        if not done:
            raise HTTPException(404, "Job not found")
        return done

    jobs.set_running(job_id)
    try:
        html = renderer.render(job["template"], job["data"])
        sender = senders.get_sender(job["sender"])
        recipients = [Recipient(email=row["email"], name=row.get("name")) for row in batch]
        result = sender.send(html, job["subject"], recipients, job["from_addr"], job["from_name"])
        sent_ids, failed_rows = _split_batch_result(batch, result.errors, result.failed)
        return jobs.mark_batch_result(job_id, sent_ids, failed_rows)
    except Exception as e:
        failed = jobs.fail_job(job_id, f"발송 실패: {e}", [str(e)])
        if not failed:
            raise HTTPException(404, "Job not found")
        return failed


@app.get("/api/send/{job_id}")
def get_send_status(job_id: str):
    job = jobs.get_job(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    return job


@app.get("/api/jobs")
def get_jobs():
    return jobs.list_jobs()


# ========== 수신거부 ==========

class UnsubIn(BaseModel):
    email: str


def _public_status_page(
    *,
    title: str,
    heading: str,
    message: str,
    email: str = "",
    status_code: int = 200,
    success: bool = True,
    footer: str = "전인교육학회 Academic Society for Human Completion",
):
    import html as _html

    icon = "✓" if success else "!"
    icon_class = "icon" if success else "icon warn"
    email_html = f'<div class="email">{_html.escape(email)}</div>' if email else ""
    return HTMLResponse(
        f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{_html.escape(title)}</title>
<style>{_PUBLIC_PAGE_CSS}</style></head><body>
  <div class="card">
    <div class="card-top">
      <div class="kicker">처리 결과</div>
      <h1>전인교육학회</h1>
      <p>요청하신 처리 결과를 확인해 주세요.</p>
    </div>
    <div class="card-body status">
      <div class="{icon_class}">{icon}</div>
      <h2>{_html.escape(heading)}</h2>
      <p>{message}</p>
      {email_html}
    </div>
    <div class="card-footer">{footer}</div>
  </div>
</body></html>""",
        status_code=status_code,
    )


@app.get("/api/unsubscribe", response_class=HTMLResponse)
def unsubscribe_landing(email: str = ""):
    """이메일에서 수신거부 링크 클릭 시 진입하는 확인 페이지"""
    import html as _html
    _email = _html.escape(email.strip())
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>전인교육학회 수신거부</title>
<style>{_PUBLIC_PAGE_CSS}</style></head><body>
  <div class="card">
    <div class="card-top">
      <div class="kicker">수신거부</div>
      <h1>전인교육학회</h1>
      <p>더 이상 소식을 원하지 않으시면 이메일 작성 후 수신거부를 눌러주세요.</p>
    </div>
    <div class="card-body">
      <form method="POST" action="/api/unsubscribe">
        <div class="form-field">
          <label>이메일 <span class="req">*</span></label>
          <input type="email" name="email" required placeholder="example@email.com" value="{_email}">
        </div>
        <button type="submit" class="btn">수신거부</button>
      </form>
    </div>
    <div class="card-footer">전인교육학회 Academic Society for Human Completion</div>
  </div>
</body></html>""")


@app.post("/api/unsubscribe", response_class=HTMLResponse)
def unsubscribe_submit(email: str = Form(...)):
    """수신거부 확인 버튼 제출 처리"""
    email = email.strip().lower()
    if not email or "@" not in email:
        return _public_status_page(
            title="수신거부 오류",
            heading="이메일 주소를 확인해 주세요",
            message="유효한 이메일 주소를 입력한 뒤 다시 시도해 주세요.",
            status_code=400,
            success=False,
        )
    try:
        added = storage.add_unsubscribed(email)
    except Exception as exc:
        print(f"unsubscribe failed for {email}: {exc}")
        return _public_status_page(
            title="수신거부 오류",
            heading="수신거부 처리에 실패했습니다",
            message="잠시 후 다시 시도해 주세요.<br>문제가 계속되면 사무국으로 연락해 주세요.",
            email=email,
            status_code=500,
            success=False,
            footer="info@humancompletion.org",
        )
    if added:
        return _public_status_page(
            title="전인교육학회 수신거부 완료",
            heading="수신거부가 완료되었습니다",
            message="아래 이메일 주소는 더 이상 전인교육학회 소식을 수신하지 않습니다.",
            email=email,
            footer="잘못 클릭하셨다면 사무국으로 연락 주세요.<br>info@humancompletion.org",
        )
    return _public_status_page(
        title="전인교육학회 수신거부 확인",
        heading="이미 수신거부된 이메일입니다",
        message="아래 이메일 주소는 이미 수신거부 명단에 등록되어 있습니다.",
        email=email,
        footer="잘못 클릭하셨다면 사무국으로 연락 주세요.<br>info@humancompletion.org",
    )


@app.get("/api/unsubscribed")
def list_unsubscribed():
    return {"emails": sorted(_load_unsubscribed_or_error())}


@app.get("/api/unsubscribed/export")
def export_unsubscribed():
    csv_str = storage.export_unsubscribed_csv()
    return Response(
        content=csv_str,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="unsubscribed.csv"'},
    )


@app.post("/api/unsubscribed")
def post_unsubscribe(body: UnsubIn):
    added = storage.add_unsubscribed(body.email)
    return {"ok": True, "added": added}


@app.delete("/api/unsubscribed/{email}")
def delete_unsubscribed(email: str):
    storage.remove_unsubscribed(email)
    return {"ok": True}


# ========== 수신동의(구독) ==========

_PUBLIC_PAGE_CSS = """
  * { box-sizing: border-box; }
  body { font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif; background:#f4f6f5; color:#0f172a; margin:0; padding:24px; display:flex; align-items:center; justify-content:center; min-height:100vh; }
  .card { background:#ffffff; border:1px solid #e3e9e5; border-radius:14px; box-shadow:0 18px 48px rgba(15,23,42,0.10); max-width:460px; width:100%; overflow:hidden; }
  .card-top { background:#1a6b4a; padding:30px 34px 25px; color:#ffffff; }
  .kicker { margin:0 0 8px; color:rgba(255,255,255,0.72); font-size:11px; font-weight:800; letter-spacing:0.14em; text-transform:uppercase; }
  .card-top h1 { margin:0; font-size:22px; font-weight:800; line-height:1.25; letter-spacing:-0.3px; }
  .card-top p { margin:8px 0 0; font-size:13px; line-height:1.65; color:rgba(255,255,255,0.86); }
  .card-body { padding:30px 34px 28px; }
  .form-field { margin-bottom:16px; }
  label { display:block; font-size:13px; font-weight:600; color:#334155; margin-bottom:6px; }
  label .req { color:#e53e3e; }
  input[type=text], input[type=email] { width:100%; min-height:42px; padding:10px 13px; border:1px solid #d9e2dc; border-radius:8px; background:#ffffff; color:#0f172a; font-size:14px; font-family:inherit; outline:none; transition:border .2s, box-shadow .2s; }
  input:focus { border-color:#1a6b4a; box-shadow:0 0 0 3px rgba(26,107,74,0.11); }
  .consent { margin:18px 0 24px; padding:15px 16px; background:#f8faf9; border-radius:8px; border:1px solid #e8efe9; }
  .consent-row { display:flex; gap:10px; align-items:flex-start; }
  .consent input[type=checkbox] { width:16px; height:16px; margin:3px 0 0; accent-color:#1a6b4a; flex-shrink:0; }
  .consent-label { display:block; margin:0; color:#334155; font-size:12.5px; font-weight:500; line-height:1.65; cursor:pointer; }
  .consent-label strong { display:block; color:#1f2937; font-size:13px; font-weight:800; }
  .consent-label span { display:block; margin-top:2px; color:#64748b; }
  .consent-detail { margin-top:12px; border-top:1px solid #e3ebe6; padding-top:10px; }
  .consent-detail summary { cursor:pointer; color:#1a6b4a; font-size:12px; font-weight:800; line-height:1.5; list-style:none; }
  .consent-detail summary::-webkit-details-marker { display:none; }
  .consent-detail summary::after { content:'보기'; float:right; color:#64748b; font-weight:700; }
  .consent-detail[open] summary::after { content:'닫기'; }
  .notice-list { margin:10px 0 0; padding:0; border:1px solid #edf2ee; border-radius:8px; overflow:hidden; background:#ffffff; }
  .notice-row { display:grid; grid-template-columns:92px 1fr; margin:0; border-top:1px solid #edf2ee; }
  .notice-row:first-child { border-top:0; }
  .notice-row dt { margin:0; padding:9px 10px; background:#f7faf8; color:#52635a; font-size:11.5px; font-weight:800; line-height:1.55; }
  .notice-row dd { margin:0; padding:9px 10px; color:#475569; font-size:11.5px; line-height:1.65; }
  .btn { width:100%; min-height:44px; padding:12px 16px; background:#1a6b4a; color:#fff; border:none; border-radius:8px; font-size:14px; font-weight:800; font-family:inherit; cursor:pointer; transition:background .2s, transform .2s; }
  .btn:hover { background:#15573d; transform:translateY(-1px); }
  .result { text-align:center; }
  .status { padding-top:32px; text-align:center; }
  .icon { width:58px; height:58px; border-radius:50%; background:#ecf5f0; color:#1a6b4a; display:inline-flex; align-items:center; justify-content:center; font-size:28px; font-weight:800; margin:0 auto 18px; }
  .icon.warn { background:#fff7ed; color:#b45309; }
  .status h2 { color:#0f172a; font-size:20px; font-weight:800; line-height:1.35; letter-spacing:-0.3px; margin:0 0 9px; }
  .status p { color:#475569; font-size:14px; line-height:1.75; margin:0 0 18px; }
  .email { display:inline-block; max-width:100%; overflow-wrap:anywhere; background:#f7f9f8; border:1px solid #dde6e0; padding:7px 14px; border-radius:8px; color:#1a6b4a; font-size:13px; font-weight:700; }
  .card-footer { border-top:1px solid #edf2ee; padding:17px 34px 22px; color:#8a9a91; font-size:11px; line-height:1.6; text-align:center; }
  @media (max-width:480px) {
    body { padding:16px; }
    .card-top { padding:28px 24px 23px; }
    .card-body { padding:28px 24px 26px; }
    .card-footer { padding:16px 24px 21px; }
  }
"""


@app.get("/api/subscribe", response_class=HTMLResponse)
def subscribe_form(email: str = "", name: str = ""):
    """수신동의 폼 페이지"""
    import html as _html
    _email = _html.escape(email.strip())
    _name = _html.escape(name.strip())
    return HTMLResponse(f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>전인교육학회 수신동의</title>
<style>{_PUBLIC_PAGE_CSS}</style></head><body>
  <div class="card">
    <div class="card-top">
      <div class="kicker">수신동의</div>
      <h1>전인교육학회</h1>
      <p>학회지, 학술대회, 캠프 등 소식을 이메일로 받아보세요.</p>
    </div>
    <div class="card-body">
      <form method="POST" action="/api/subscribe">
        <div class="form-field">
          <label>이름 <span class="req">*</span></label>
          <input type="text" name="name" required placeholder="홍길동" value="{_name}">
        </div>
        <div class="form-field">
          <label>이메일 <span class="req">*</span></label>
          <input type="email" name="email" required placeholder="example@email.com" value="{_email}">
        </div>
        <div class="form-field">
          <label>소속</label>
          <input type="text" name="organization" placeholder="대학교/기관명 (선택)">
        </div>
        <div class="consent">
          <div class="consent-row">
            <input type="checkbox" id="agree" name="agree" required>
            <label for="agree" class="consent-label">
              <strong>개인정보 수집 및 이용에 동의합니다.</strong>
              <span>뉴스레터 발송을 위한 필수 동의입니다.</span>
            </label>
          </div>
          <details class="consent-detail">
            <summary>개인정보 수집 및 이용 안내</summary>
            <dl class="notice-list">
              <div class="notice-row">
                <dt>수집 목적</dt>
                <dd>전인교육학회 뉴스레터, 학회지, 학술대회, 캠프 등 소식 및 안내 이메일 발송</dd>
              </div>
              <div class="notice-row">
                <dt>수집 항목</dt>
                <dd>필수: 이름, 이메일 / 선택: 소속</dd>
              </div>
              <div class="notice-row">
                <dt>보유 기간</dt>
                <dd>수신동의 철회 또는 수신거부 시까지 보관합니다. 수신거부 시 발송 명단에서는 제외되며, 재발송 방지를 위해 이메일 주소는 수신거부 명단으로 보관됩니다.</dd>
              </div>
              <div class="notice-row">
                <dt>동의 거부</dt>
                <dd>동의를 거부할 권리가 있으며, 거부 시 뉴스레터 및 관련 소식 이메일을 받을 수 없습니다.</dd>
              </div>
            </dl>
          </details>
        </div>
        <button type="submit" class="btn">수신동의</button>
      </form>
    </div>
    <div class="card-footer">전인교육학회 Academic Society for Human Completion</div>
  </div>
</body></html>""")


@app.post("/api/subscribe", response_class=HTMLResponse)
def subscribe_submit(
    request: Request,
    email: str = Form(...),
    name: str = Form(""),
    organization: str = Form(""),
    agree: str = Form(""),
):
    """수신동의 폼 제출 처리"""
    email = email.strip().lower()
    if not email or "@" not in email:
        return _public_status_page(
            title="수신동의 오류",
            heading="이메일 주소를 확인해 주세요",
            message="유효한 이메일 주소를 입력한 뒤 다시 시도해 주세요.",
            status_code=400,
            success=False,
        )
    try:
        forwarded_for = request.headers.get("x-forwarded-for", "")
        ip = forwarded_for.split(",", 1)[0].strip() or (request.client.host if request.client else "")
        added = storage.add_subscriber(
            email,
            name,
            organization,
            consent_source="public_subscribe_form",
            consent_version=CONSENT_VERSION,
            ip=ip,
            user_agent=request.headers.get("user-agent", ""),
        )
    except Exception as exc:
        print(f"subscribe failed for {email}: {exc}")
        return _public_status_page(
            title="수신동의 오류",
            heading="수신동의 처리에 실패했습니다",
            message="잠시 후 다시 시도해 주세요.<br>문제가 계속되면 사무국으로 연락해 주세요.",
            email=email,
            status_code=500,
            success=False,
            footer="info@humancompletion.org",
        )
    if added:
        msg_title = "수신동의가 완료되었습니다"
        msg_body = "앞으로 전인교육학회 뉴스레터를<br>이메일로 받아보실 수 있습니다."
    else:
        msg_title = "이미 등록된 이메일입니다"
        msg_body = "해당 이메일은 이미 수신동의 되어 있습니다.<br>추가 조치가 필요하지 않습니다."
    return _public_status_page(
        title="수신동의 완료",
        heading=msg_title,
        message=msg_body,
        email=email,
    )


@app.get("/api/subscribers")
def list_subscribers():
    subs = storage.load_subscribers()
    return {"subscribers": subs, "count": len(subs)}


@app.delete("/api/subscribers/{email:path}")
def delete_subscriber(email: str):
    storage.remove_subscriber(email)
    return {"ok": True}


@app.get("/api/subscribers/export")
def export_subscribers():
    csv_str = storage.export_subscribers_csv()
    return Response(
        content=csv_str,
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="subscribers.csv"'},
    )


@app.get("/api/manual")
def get_manual():
    import os
    candidates = [
        os.path.join(os.path.dirname(__file__), "..", "MANUAL.md"),
        "/app/MANUAL.md",
        "MANUAL.md",
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                return {"content": f.read()}
    return {"content": "# 매뉴얼 파일을 찾을 수 없습니다."}


@app.get("/health")
def health():
    return {
        "ok": True,
        "app_env": config.APP_ENV,
        "vercel_runtime": config.RUNNING_ON_VERCEL,
        "requires_supabase": config.REQUIRE_SUPABASE,
        "storage": "supabase" if storage.using_supabase() else "local",
    }


@app.get("/{full_path:path}", include_in_schema=False)
def serve_frontend(full_path: str):
    if full_path.startswith("api/"):
        raise HTTPException(404, "Not found")

    dist = Path(config.FRONTEND_DIST_DIR)
    index = dist / "index.html"
    if not index.exists():
        raise HTTPException(404, "Frontend build not found")

    target = (dist / full_path).resolve()
    try:
        target.relative_to(dist.resolve())
    except ValueError:
        raise HTTPException(404, "Not found")

    if full_path and target.is_file():
        return FileResponse(target)
    return FileResponse(index)
